import openpyxl
import shutil
import os
import sys

# Force the console to use UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# Directory containing Excel files (script directory)
excel_folder = os.getcwd()
text_file = "APK Version.txt"

# Read the text file and store mappings in a dictionary
mappings = {}

with open(text_file, "r", encoding="utf-8") as file:
    for line in file:
        if ":" in line:
            # Clean the data
            key, value = line.split(":")
            key = key.strip().replace("(", "").replace(")", "")  # Remove parentheses and spaces
            mappings[key] = value.strip()

# Verify that the dictionary is populated
print("\nMappings found in APK Version.txt:")
for k, v in mappings.items():
    print(f"{k} → {v}")

# Iterate through all Excel files in the folder
for file in os.listdir(excel_folder):
    if file.endswith(".xlsx") or file.endswith(".xlsm"):
        file_path = os.path.join(excel_folder, file)
        temp_file = os.path.join(excel_folder, "temp_" + file)

        # Create a temporary copy
        shutil.copy(file_path, temp_file)

        # Load the Excel file
        wb = openpyxl.load_workbook(temp_file)
        ws = wb.active  # Select the active sheet (ensure this is correct)

        # Read the value from cell U4
        value_u4 = ws["U4"].value
        if value_u4:
            value_u4 = value_u4.strip()  # Remove invisible spaces

        print(f"\nFile: {file.encode('utf-8', 'ignore').decode()}")  # Handle encoding
        print("Value in cell U4:", value_u4)

        # Check if U4 value exists in the text file
        if value_u4 in mappings:
            mapped_value = mappings[value_u4]
            print("Value found in the text file:", mapped_value)

            # Update cell V4
            ws["V4"] = mapped_value

            # Save the modified Excel file
            wb.save(temp_file)
            print("Update of V4 completed.")
        else:
            print("No match found for U4 in APK Version.txt.")

        # Close the file
        wb.close()

        # Replace the original file with the updated version
        shutil.move(temp_file, file_path)

print("\nProcessing completed for all files.")
